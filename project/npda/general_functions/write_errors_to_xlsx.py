# import types
import io

# import third-party libaries
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas.api.types import is_datetime64_any_dtype

# import functions
from project.npda.general_functions.csv.csv_parse import csv_parse

# import csv mappings
from ...constants.csv_headings import csv_definition_for


def write_errors_to_xlsx(
    errors: dict[str, dict[str, list[str]]],
    original_csv_file_bytes: bytes,
    dataset_year: int = 2021,
) -> bytes:
    """
    Write errors to an Excel file. Highlight invalid cells in the source CSV.

    Args:
      errors A nested dictionary containing errors grouped by row index, then field.

    """

    xlsx_file = io.BytesIO()

    # Get original data
    parsed_csv = csv_parse(
        io.BytesIO(initial_bytes=original_csv_file_bytes), dataset_year=dataset_year
    )

    df = parsed_csv.df

    # Convert datetime columns to date (strip time)
    df = strip_time_in_dataframe(df)

    # Write an xlsx of the original data.
    df.to_excel(xlsx_file, sheet_name="Uploaded data (raw)", index=False)

    # If the csv file is from Jersey, add the Jersey unique identifier to the CSV headings, otherwise add the England unique identifier
    df_errors = flatten_errors(
        errors=errors,
        original_data=df,
        identifier_column=parsed_csv.identifier_column,
    )

    # Add sheet that lists the errors.
    with pd.ExcelWriter(xlsx_file, mode="a", engine="openpyxl") as writer:
        df_errors.to_excel(writer, sheet_name="Errors - Overview", index=False)

    # Load the workbook in openpyxl
    wb: Workbook = load_workbook(xlsx_file)

    # Set text to red
    overview_sheet = wb["Errors - Overview"]
    for row in overview_sheet.iter_rows(min_row=2, min_col=4):
        for cell in row:
            cell.font = Font(color="FF0000")

    # Setup the styled worksheet
    styled_sheet: Worksheet = wb.copy_worksheet(wb["Uploaded data (raw)"])
    styled_sheet.title = (
        "Uploaded data (comments)"  # You can set any name for the copied sheet
    )

    # Style the openpyxl worksheet to highlight in red erroneous/invalid cells.
    # Also add comments to annotate the actual error.
    for _, patient_errors in df_errors.iterrows():
        row_index = patient_errors["Original CSV Row"]

        field_name = patient_errors["Column"]
        field_errors = patient_errors["Errors"]

        column_index = find_column_index_by_name(field_name, styled_sheet)
        if column_index:
            styled_sheet.cell(row=row_index, column=column_index).fill = PatternFill(
                patternType="solid", fgColor="FFC9C9"
            )  # Change color to red.
            styled_sheet.cell(row=row_index, column=column_index).comment = Comment(
                field_errors,
                "Data Validator [Automated: RCPCH]",
                height=300,
                width=300,
            )

    # Specify the desired order by reordering the `workbook.worksheets` list
    wb._sheets = [
        wb["Uploaded data (raw)"],
        wb["Uploaded data (comments)"],
        wb["Errors - Overview"],
    ]

    # Auto-size columns to widest word (max 30 chars) and enable wrapping
    for sheet_name in [
        "Uploaded data (raw)",
        "Uploaded data (comments)",
        "Errors - Overview",
    ]:
        set_column_widths_and_wrapping(wb[sheet_name], max_width=30)

    # Save the styled sheet.
    xlsx_file = io.BytesIO()
    wb.save(xlsx_file)

    return xlsx_file.getvalue()


def find_column_index_by_name(column_name: str, ws: Worksheet) -> int | None:
    column_index = None
    for col in ws.iter_cols(
        1, ws.max_column, 1, 1
    ):  # Check headers in the first row only
        if col[0].value == column_name:
            column_index = col[0].column  # Get the column index
            break
    return column_index


def set_column_widths_and_wrapping(ws: Worksheet, max_width: int = 50) -> None:
    """
    Sets each column width to the length of the widest word found in that column,
    capped at max_width. Enables wrap text for all populated cells so longer content wraps.
    """
    # Compute width by widest word (header included)
    for col_idx in range(1, ws.max_column + 1):
        max_word_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None:
                continue
            s = str(val).replace("\n", " ")
            words = s.split()
            if words:
                max_word_len = max(max_word_len, max(len(w) for w in words))
        width = min(max_word_len if max_word_len > 0 else 10, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Enable wrapping for all populated cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(wrapText=True)


def strip_time_in_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert any datetime64 columns to date (removing time component).
    """
    df = df.copy()
    for col in df.columns:
        if is_datetime64_any_dtype(df[col]):
            # For true date-only cells in Excel, convert to Python date objects:
            df[col] = pd.to_datetime(df[col]).dt.date
    return df


def flatten_errors(
    #  {row_number: {field_name: [error_messages]}}
    errors: dict[int, dict[str, list[str]]],
    original_data: pd.DataFrame,
    identifier_column: str,
) -> pd.DataFrame:
    rows = []

    for row_ix, row_errors in errors.items():
        for field, errors in row_errors.items():
            # __all__ errors should be attached to the first column
            csv_definition = csv_definition_for(field)
            column = csv_definition["heading"] if csv_definition else identifier_column

            rows.append(
                {
                    # 0 based indexing and the column header. So + 2
                    "Original CSV Row": int(row_ix) + 2,
                    identifier_column: original_data.loc[
                        int(row_ix), identifier_column
                    ],
                    "Column": column,
                    "Errors": "; ".join(errors),
                }
            )

    return pd.DataFrame(
        rows, columns=["Original CSV Row", identifier_column, "Column", "Errors"]
    )
